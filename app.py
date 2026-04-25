# app.py
# Purpose: Streamlit app — MultiTaskOmicsNet inference
#          Image only → Tumor/Normal + Genomic Profile + Diffusion Generation
# Project: Omics-Guided Histopathology Analysis
# Author : Ranjith Kumar

import os
import io
import re
import numpy as np
import pandas as pd
from PIL import Image
import torch
import torch.nn as nn
import torch.nn.functional as F
from torchvision.models import efficientnet_b0
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import cv2
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import streamlit as st

# ──────────────────────────────────────────────
# PATHS
# ──────────────────────────────────────────────
BASE_DIR     = os.path.dirname(os.path.abspath(__file__))
MODEL_PATH   = os.path.join(BASE_DIR, "models",   "multitask_omics_best.pth")
DIFF_PATH    = os.path.join(BASE_DIR, "diffusion","diffusion_model.pth")
SAMPLES_DIR  = os.path.join(BASE_DIR, "test_samples")

GENOMIC_FEATURES = [
    "BRCA1_mutation","TP53_mutation","HER2_amplification",
    "PIK3CA_mutation","CDH1_mutation","genomic_risk_score"
]
GENOMIC_LABELS = ["BRCA1","TP53","HER2","PIK3CA","CDH1","Risk Score"]

PCAM_MEAN = np.array([0.701, 0.538, 0.692])
PCAM_STD  = np.array([0.235, 0.277, 0.213])

IMAGE_FEAT_DIM = 256
FUSION_DIM     = 128
NUM_CLASSES    = 2
GENOMIC_DIM    = 6
DROPOUT        = 0.3
COND_DIM       = 64
IMAGE_SIZE     = 32
CHANNELS       = 3
T_STEPS        = 100


# ──────────────────────────────────────────────
# MULTI-TASK CLASSIFIER
# ──────────────────────────────────────────────
class ImageEncoder(nn.Module):
    def __init__(self):
        super().__init__()
        # Load architecture only — weights loaded from our trained .pth file
        base = efficientnet_b0(weights=None)
        self.features  = base.features
        self.avgpool   = base.avgpool
        self.projector = nn.Sequential(
            nn.Dropout(p=DROPOUT),
            nn.Linear(1280, IMAGE_FEAT_DIM),
            nn.BatchNorm1d(IMAGE_FEAT_DIM),
            nn.ReLU(inplace=True),
        )
    def forward(self, x):
        x = self.features(x)
        x = self.avgpool(x)
        x = torch.flatten(x, 1)
        return self.projector(x)

class ClassificationHead(nn.Module):
    def __init__(self):
        super().__init__()
        self.head = nn.Sequential(
            nn.Linear(IMAGE_FEAT_DIM, FUSION_DIM),
            nn.BatchNorm1d(FUSION_DIM),
            nn.ReLU(inplace=True),
            nn.Dropout(p=DROPOUT),
            nn.Linear(FUSION_DIM, NUM_CLASSES),
        )
    def forward(self, x): return self.head(x)

class GenomicRegressionHead(nn.Module):
    def __init__(self):
        super().__init__()
        self.head = nn.Sequential(
            nn.Linear(IMAGE_FEAT_DIM, FUSION_DIM),
            nn.BatchNorm1d(FUSION_DIM),
            nn.ReLU(inplace=True),
            nn.Dropout(p=DROPOUT),
            nn.Linear(FUSION_DIM, GENOMIC_DIM),
            nn.Sigmoid()
        )
    def forward(self, x): return self.head(x)

class MultiTaskOmicsNet(nn.Module):
    def __init__(self):
        super().__init__()
        self.image_encoder = ImageEncoder()
        self.cls_head      = ClassificationHead()
        self.genomic_head  = GenomicRegressionHead()
    def forward(self, images):
        features     = self.image_encoder(images)
        cls_logits   = self.cls_head(features)
        genomic_pred = self.genomic_head(features)
        return cls_logits, genomic_pred


# ──────────────────────────────────────────────
# DIFFUSION MODEL
# ──────────────────────────────────────────────
class SinusoidalTimeEmbedding(nn.Module):
    def __init__(self, dim):
        super().__init__()
        self.dim = dim
    def forward(self, t):
        device = t.device
        half   = self.dim // 2
        freqs  = torch.exp(
            -torch.log(torch.tensor(10000.0)) *
            torch.arange(half, device=device) / half
        )
        args = t[:, None].float() * freqs[None]
        return torch.cat([args.sin(), args.cos()], dim=-1)

class ResidualBlock(nn.Module):
    def __init__(self, in_ch, out_ch, time_dim, cond_dim):
        super().__init__()
        self.conv1    = nn.Conv2d(in_ch,  out_ch, 3, padding=1)
        self.conv2    = nn.Conv2d(out_ch, out_ch, 3, padding=1)
        self.bn1      = nn.GroupNorm(8, out_ch)
        self.bn2      = nn.GroupNorm(8, out_ch)
        self.time_mlp = nn.Linear(time_dim, out_ch)
        self.cond_mlp = nn.Linear(cond_dim, out_ch)
        self.skip     = nn.Conv2d(in_ch, out_ch, 1) if in_ch != out_ch else nn.Identity()
    def forward(self, x, t_emb, c_emb):
        h = F.silu(self.bn1(self.conv1(x)))
        h = h + self.time_mlp(t_emb)[:, :, None, None]
        h = h + self.cond_mlp(c_emb)[:, :, None, None]
        h = F.silu(self.bn2(self.conv2(h)))
        return h + self.skip(x)

class DownBlock(nn.Module):
    def __init__(self, in_ch, out_ch, time_dim, cond_dim):
        super().__init__()
        self.res  = ResidualBlock(in_ch, out_ch, time_dim, cond_dim)
        self.down = nn.Conv2d(out_ch, out_ch, 3, stride=2, padding=1)
    def forward(self, x, t_emb, c_emb):
        x = self.res(x, t_emb, c_emb)
        return self.down(x), x

class UpBlock(nn.Module):
    def __init__(self, in_ch, skip_ch, out_ch, time_dim, cond_dim):
        super().__init__()
        self.up  = nn.ConvTranspose2d(in_ch, in_ch, 2, stride=2)
        self.res = ResidualBlock(in_ch + skip_ch, out_ch, time_dim, cond_dim)
    def forward(self, x, skip, t_emb, c_emb):
        x = self.up(x)
        x = torch.cat([x, skip], dim=1)
        return self.res(x, t_emb, c_emb)

class ConditionalUNet(nn.Module):
    def __init__(self):
        super().__init__()
        TIME_DIM = 128
        self.time_embed = nn.Sequential(
            SinusoidalTimeEmbedding(TIME_DIM),
            nn.Linear(TIME_DIM, TIME_DIM), nn.SiLU(),
            nn.Linear(TIME_DIM, TIME_DIM),
        )
        self.cond_embed = nn.Sequential(
            nn.Linear(GENOMIC_DIM, 32), nn.SiLU(),
            nn.Linear(32, COND_DIM),
        )
        self.enc1       = DownBlock(CHANNELS, 32,  TIME_DIM, COND_DIM)
        self.enc2       = DownBlock(32,       64,  TIME_DIM, COND_DIM)
        self.enc3       = DownBlock(64,       128, TIME_DIM, COND_DIM)
        self.bottleneck = ResidualBlock(128, 128, TIME_DIM, COND_DIM)
        self.dec3       = UpBlock(128, 128, 64, TIME_DIM, COND_DIM)
        self.dec2       = UpBlock( 64,  64, 32, TIME_DIM, COND_DIM)
        self.dec1       = UpBlock( 32,  32, 32, TIME_DIM, COND_DIM)
        self.out        = nn.Conv2d(32, CHANNELS, 1)
    def forward(self, x, t, genomic):
        t_emb = self.time_embed(t)
        c_emb = self.cond_embed(genomic)
        x, s1 = self.enc1(x, t_emb, c_emb)
        x, s2 = self.enc2(x, t_emb, c_emb)
        x, s3 = self.enc3(x, t_emb, c_emb)
        x = self.bottleneck(x, t_emb, c_emb)
        x = self.dec3(x, s3, t_emb, c_emb)
        x = self.dec2(x, s2, t_emb, c_emb)
        x = self.dec1(x, s1, t_emb, c_emb)
        return self.out(x)


# ──────────────────────────────────────────────
# GRAD-CAM
# ──────────────────────────────────────────────
class GradCAM:
    def __init__(self, model):
        self.model       = model
        self.gradients   = None
        self.activations = None
        target = model.image_encoder.features[-1]
        target.register_forward_hook(self._save_activation)
        target.register_full_backward_hook(self._save_gradient)
    def _save_activation(self, module, input, output):
        self.activations = output.detach()
    def _save_gradient(self, module, grad_input, grad_output):
        self.gradients = grad_output[0].detach()
    def generate(self, img_tensor, class_idx=1):
        self.model.eval()
        img_tensor = img_tensor.unsqueeze(0).requires_grad_(True)
        cls_logits, _ = self.model(img_tensor)
        self.model.zero_grad()
        cls_logits[0, class_idx].backward()
        weights = self.gradients.mean(dim=[2,3], keepdim=True)
        cam     = (weights * self.activations).sum(dim=1, keepdim=True)
        cam     = torch.relu(cam).squeeze().numpy()
        cam     = cv2.resize(cam, (96,96))
        cam     = (cam - cam.min()) / (cam.max() - cam.min() + 1e-8)
        return cam


# ──────────────────────────────────────────────
# UTILITIES
# ──────────────────────────────────────────────
@st.cache_resource
def load_classifier():
    model = MultiTaskOmicsNet()
    model.load_state_dict(torch.load(MODEL_PATH, map_location="cpu"))
    model.eval()
    return model

@st.cache_resource
def load_diffusion():
    ckpt       = torch.load(DIFF_PATH, map_location="cpu")
    diff_model = ConditionalUNet()
    diff_model.load_state_dict(ckpt["model_state_dict"])
    diff_model.eval()
    return diff_model, ckpt["betas"], ckpt["alphas"], ckpt["alpha_bars"]

@st.cache_data
def get_sample_files():
    if not os.path.exists(SAMPLES_DIR):
        return []
    return sorted([f for f in os.listdir(SAMPLES_DIR)
                   if f.endswith((".png",".jpg",".jpeg"))])

def extract_idx(filename):
    m = re.search(r"idx(\d+)", filename)
    return int(m.group(1)) if m else -1

def enhance_image(pil_img):
    img  = np.array(pil_img.convert("RGB"))
    lab  = cv2.cvtColor(img, cv2.COLOR_RGB2LAB)
    l, a, b = cv2.split(lab)
    clahe   = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8,8))
    lab_eq  = cv2.merge([clahe.apply(l), a, b])
    enhanced  = cv2.cvtColor(lab_eq, cv2.COLOR_LAB2RGB)
    sharpened = cv2.filter2D(enhanced, -1, np.array([[0,-1,0],[-1,5,-1],[0,-1,0]]))
    return Image.fromarray(sharpened)

def preprocess(pil_img):
    img    = pil_img.convert("RGB").resize((96,96))
    img_np = np.array(img, dtype=np.float32) / 255.0
    norm   = (img_np - PCAM_MEAN) / PCAM_STD
    tensor = torch.tensor(norm.transpose(2,0,1), dtype=torch.float32)
    return tensor, img_np

@torch.no_grad()
def generate_synthetic(diff_model, genomic_vec, betas, alphas, alpha_bars,
                       ddim_steps=50, eta=0.0):
    """
    DDIM sampler — produces sharper images than DDPM.
    ddim_steps: number of inference steps (50 is enough, vs 100 for DDPM)
    eta=0.0: deterministic sampling (sharpest output)
    eta=1.0: stochastic (same as DDPM)
    """
    T   = T_STEPS
    g   = genomic_vec.unsqueeze(0)
    x   = torch.randn(1, CHANNELS, IMAGE_SIZE, IMAGE_SIZE)

    # Build DDIM timestep sequence — evenly spaced subset of [0, T]
    step_size = T // ddim_steps
    timesteps = list(reversed(range(0, T, step_size)))  # e.g. [99,97,95,...,1]

    for i, t_idx in enumerate(timesteps):
        t_tensor   = torch.tensor([t_idx], dtype=torch.long)
        pred_noise = diff_model(x, t_tensor, g)

        alpha_bar_t = alpha_bars[t_idx]

        # Predict x0
        x0_pred = (x - torch.sqrt(1 - alpha_bar_t) * pred_noise) / torch.sqrt(alpha_bar_t)
        x0_pred = x0_pred.clamp(-1, 1)

        if i < len(timesteps) - 1:
            t_prev         = timesteps[i + 1]
            alpha_bar_prev = alpha_bars[t_prev]
        else:
            alpha_bar_prev = torch.tensor(1.0)

        # DDIM update
        sigma = (eta *
                 torch.sqrt((1 - alpha_bar_prev) / (1 - alpha_bar_t)) *
                 torch.sqrt(1 - alpha_bar_t / alpha_bar_prev))

        direction = torch.sqrt(1 - alpha_bar_prev - sigma**2) * pred_noise
        noise     = sigma * torch.randn_like(x) if eta > 0 else 0

        x = torch.sqrt(alpha_bar_prev) * x0_pred + direction + noise

    img = (x.clamp(-1, 1) + 1) / 2.0
    img = F.interpolate(img, size=(96, 96), mode="bicubic", align_corners=False)
    return img.squeeze(0).permute(1, 2, 0).numpy()

def make_excel(image_idx, filename, label, confidence, genomic_pred):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Prediction Report"
    hdr_fill = PatternFill("solid", start_color="1F4E79")
    hdr_font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    center   = Alignment(horizontal="center", vertical="center")
    thin     = Border(left=Side(style="thin"), right=Side(style="thin"),
                      top=Side(style="thin"),  bottom=Side(style="thin"))
    res_fill = PatternFill("solid",
               start_color="FFE0E0" if label=="TUMOR" else "E0F0E0")
    rows = [
        ["Field",              "Predicted Value"],
        ["Filename",           filename],
        ["Image Index",        image_idx],
        ["Prediction",         label],
        ["Confidence",         f"{confidence:.2%}"],
        ["─── Predicted Omics ───", "─────────────────"],
        ["BRCA1 Mutation",     round(float(genomic_pred[0]),4)],
        ["TP53 Mutation",      round(float(genomic_pred[1]),4)],
        ["HER2 Amplification", round(float(genomic_pred[2]),4)],
        ["PIK3CA Mutation",    round(float(genomic_pred[3]),4)],
        ["CDH1 Mutation",      round(float(genomic_pred[4]),4)],
        ["Genomic Risk Score", round(float(genomic_pred[5]),4)],
        ["─── Note ───",       "Genomic values predicted from image by MultiTaskOmicsNet"],
    ]
    for r, row in enumerate(rows, 1):
        for c, val in enumerate(row, 1):
            cell = ws.cell(r, c, value=val)
            cell.alignment = center; cell.border = thin
            if r == 1:
                cell.font = hdr_font; cell.fill = hdr_fill
            else:
                cell.font = Font(name="Arial", size=10)
                cell.fill = res_fill
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 45
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return buf


# ──────────────────────────────────────────────
# MAIN APP
# ──────────────────────────────────────────────
def main():
    st.set_page_config(
        page_title="OmicsGuidedNet",
        page_icon="🔬",
        layout="wide"
    )

    st.markdown("""
    <h1 style='text-align:center;color:#1F4E79;'>🔬 OmicsGuidedNet</h1>
    <p style='text-align:center;color:#888;font-size:15px;'>
        Multi-Task Biomarker Discovery — Image → Tumor/Normal + Genomic Profile
    </p>
    <hr>
    """, unsafe_allow_html=True)

    # Load models
    with st.spinner("Loading models..."):
        classifier                        = load_classifier()
        gradcam                           = GradCAM(classifier)
        diff_model, betas, alphas, alpha_bars = load_diffusion()
        samples                           = get_sample_files()

    # Sidebar
    st.sidebar.header("⚙️ Settings")
    enhance  = st.sidebar.checkbox("Enhance image (CLAHE + Sharpen)", value=True)
    gen_diff = st.sidebar.checkbox("Generate diffusion patch",         value=True)
    st.sidebar.markdown("---")
    st.sidebar.markdown("**Classifier:** MultiTaskOmicsNet")
    st.sidebar.markdown("**Task 1:** Tumor / Normal prediction")
    st.sidebar.markdown("**Task 2:** Genomic profile prediction")
    st.sidebar.markdown("**Diffusion:** Conditional DDPM (100 steps)")
    st.sidebar.markdown("**Test AUC:** 0.9467 | **Genomic MAE:** 0.3950")

    # ── STEP 1: Select image ──
    st.subheader("Step 1 — Select Preloaded Test Image")

    if not samples:
        st.warning(f"No samples found in {SAMPLES_DIR}")
        return

    selected   = st.selectbox(
        "Choose a test patch",
        samples,
        format_func=lambda f: (
            f"🔴 Tumor  — {f}" if f.startswith("tumor") else f"🟢 Normal — {f}"
        )
    )
    pil_img    = Image.open(os.path.join(SAMPLES_DIR, selected))
    image_idx  = extract_idx(selected)
    true_label = "TUMOR" if selected.startswith("tumor") else "NORMAL"
    enh_img    = enhance_image(pil_img) if enhance else pil_img

    # Show original vs enhanced
    col_o, col_e = st.columns(2)
    with col_o:
        st.image(pil_img, width=160, caption="Original patch")
    with col_e:
        st.image(enh_img, width=160,
                 caption="Enhanced (CLAHE + Sharpen)" if enhance else "No enhancement")

    true_color = "#e74c3c" if true_label=="TUMOR" else "#27ae60"
    st.markdown(
        f"**Ground truth:** "
        f"<span style='background:{true_color};color:white;"
        f"padding:3px 12px;border-radius:4px;'>{true_label}</span> &nbsp;"
        f"**Index:** `{image_idx}`",
        unsafe_allow_html=True
    )

    # ── STEP 2: Predict ──
    st.markdown("---")
    st.subheader("Step 2 — Run Analysis")

    if st.button("🔍 Analyze", type="primary", use_container_width=True):

        img_tensor, img_np = preprocess(enh_img)

        # Multi-task inference — image only
        classifier.eval()
        with torch.no_grad():
            cls_logits, genomic_pred = classifier(img_tensor.unsqueeze(0))
            probs      = torch.softmax(cls_logits, dim=1)[0]
            pred_class = cls_logits.argmax(dim=1).item()
            confidence = probs[pred_class].item()
            genomic_vals = genomic_pred.squeeze(0).numpy()

        label = "TUMOR" if pred_class == 1 else "NORMAL"

        # ── Prediction result ──
        color = "#e74c3c" if label=="TUMOR" else "#27ae60"
        st.markdown(f"""
        <div style='background:{color};padding:20px;border-radius:10px;
                    text-align:center;margin:10px 0;'>
            <h2 style='color:white;margin:0;'>{label}</h2>
            <p style='color:white;margin:4px 0 0;font-size:18px;'>
                Confidence: {confidence:.2%}
            </p>
        </div>
        """, unsafe_allow_html=True)

        c1, c2 = st.columns(2)
        with c1:
            st.metric("Normal", f"{probs[0].item():.2%}")
            st.progress(float(probs[0].item()))
        with c2:
            st.metric("Tumor", f"{probs[1].item():.2%}")
            st.progress(float(probs[1].item()))

        st.markdown("---")

        # ── Grad-CAM ──
        st.subheader("🔥 Grad-CAM — Region Focus")
        cam = gradcam.generate(img_tensor, class_idx=pred_class)
        fig, axes = plt.subplots(1, 3, figsize=(10,3))
        axes[0].imshow(np.clip(img_np,0,1)); axes[0].set_title("Original"); axes[0].axis("off")
        axes[1].imshow(cam, cmap="jet");     axes[1].set_title("Grad-CAM"); axes[1].axis("off")
        axes[2].imshow(np.clip(img_np,0,1))
        axes[2].imshow(cam, cmap="jet", alpha=0.45)
        axes[2].set_title("Overlay"); axes[2].axis("off")
        plt.tight_layout()
        st.pyplot(fig); plt.close()

        st.markdown("---")

        # ── Diffusion model ──
        st.markdown("""
        <div style='background:#0d1b2a;border:2px solid #00d4ff;
                    border-radius:12px;padding:16px;margin:10px 0;'>
            <h3 style='color:#00d4ff;margin:0 0 4px 0;'>
                🧬 Diffusion Model — Synthetic Patch Generation
            </h3>
            <p style='color:#aaa;margin:0;font-size:13px;'>
                Conditional DDPM · 100 denoising timesteps ·
                Conditioned on predicted genomic profile · 32×32 → 96×96
            </p>
        </div>
        """, unsafe_allow_html=True)

        if gen_diff:
            with st.spinner("Running diffusion model (100 timesteps)..."):
                genomic_tensor = torch.tensor(genomic_vals, dtype=torch.float32)
                syn_patch = generate_synthetic(
                    diff_model, genomic_tensor, betas, alphas, alpha_bars
                )

            col_r, col_g = st.columns(2)
            with col_r:
                st.image(np.clip(img_np,0,1),
                         caption="Real patch (input)",
                         width=300)
                st.markdown(
                    "<p style='text-align:center;color:#888;font-size:12px;'>"
                    "Source: PCam test set</p>",
                    unsafe_allow_html=True
                )
            with col_g:
                st.image(np.clip(syn_patch,0,1),
                         caption="🧬 Diffusion-generated patch",
                         width=300)
                st.markdown(
                    "<p style='text-align:center;color:#00d4ff;font-size:12px;'>"
                    "Generated by Conditional DDPM<br>"
                    "conditioned on predicted genomic profile</p>",
                    unsafe_allow_html=True
                )

            st.markdown("""
            <div style='background:#1a1a2e;border-left:4px solid #00d4ff;
                        padding:10px 14px;border-radius:6px;margin-top:8px;'>
                <p style='color:#ccc;margin:0;font-size:13px;'>
                    <b style='color:#00d4ff;'>How this works:</b>
                    The MultiTaskOmicsNet first predicts the genomic profile
                    from the image. The diffusion model then uses this predicted
                    profile as a condition to generate a synthetic patch —
                    showing what tissue with that genomic signature looks like.
                </p>
            </div>
            """, unsafe_allow_html=True)
        else:
            st.info("Enable 'Generate diffusion patch' in sidebar to see this.")

        st.markdown("---")

        # ── Predicted Omics Profile ──
        st.subheader("🧬 Predicted Omics Profile")
        st.caption("Predicted directly from image by MultiTaskOmicsNet — no database lookup")

        risk       = float(genomic_vals[5])
        risk_label = "High" if risk > 0.6 else "Medium" if risk > 0.3 else "Low"
        risk_color = "#e74c3c" if risk > 0.6 else "#f39c12" if risk > 0.3 else "#27ae60"

        st.markdown(f"""
        <div style='background:#f8f9fa;padding:12px;border-radius:8px;
                    border-left:4px solid {risk_color};margin-bottom:14px;'>
            <b>Predicted Genomic Risk:
            <span style='color:{risk_color}'>{risk_label} ({risk:.3f})</span></b>
        </div>
        """, unsafe_allow_html=True)

        # Mutation badges with actual predicted values
        mut_cols = st.columns(5)
        for i, (feat, display) in enumerate(zip(
            ["BRCA1_mutation","TP53_mutation","HER2_amplification",
             "PIK3CA_mutation","CDH1_mutation"],
            ["BRCA1","TP53","HER2","PIK3CA","CDH1"]
        )):
            raw_val  = float(genomic_vals[i])
            bin_val  = 1 if raw_val >= 0.5 else 0
            mcolor   = "#e74c3c" if bin_val == 1 else "#27ae60"
            mut_cols[i].markdown(
                f"<div style='text-align:center;background:{mcolor};"
                f"color:white;padding:10px 4px;border-radius:8px;'>"
                f"<b style='font-size:13px;'>{display}</b><br>"
                f"<span style='font-size:11px;'>{'Mutated' if bin_val==1 else 'Normal'}</span><br>"
                f"<b style='font-size:16px;'>{raw_val:.3f}</b></div>",
                unsafe_allow_html=True
            )

        # Full genomic table
        st.markdown("<br>", unsafe_allow_html=True)
        gdf = pd.DataFrame([{
            "Feature"        : lbl,
            "Predicted Value": round(float(genomic_vals[i]), 4),
            "Binary"         : str(1 if (i < 5 and genomic_vals[i] >= 0.5) else
                                   ("N/A" if i == 5 else 0)),
            "Status"         : ("Mutated" if (i < 5 and genomic_vals[i] >= 0.5)
                                else ("High"   if (i==5 and genomic_vals[i] > 0.6)
                                      else ("Medium" if (i==5 and genomic_vals[i] > 0.3)
                                            else ("Low" if i==5 else "Normal"))))
        } for i, lbl in enumerate(GENOMIC_LABELS)])
        st.dataframe(gdf, use_container_width=True, hide_index=True)

        st.markdown("---")

        # ── Excel download ──
        excel_buf = make_excel(image_idx, selected, label, confidence, genomic_vals)
        st.download_button(
            label="📥 Download Excel Report",
            data=excel_buf,
            file_name=f"omicsguided_{selected}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )


if __name__ == "__main__":
    main()